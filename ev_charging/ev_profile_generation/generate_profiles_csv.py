import warnings
warnings.simplefilter(action='ignore', category=FutureWarning)

import os
import sys
import csv
import time
import openpyxl
from copy import copy
from datetime import datetime, timedelta

from datafev.routines.scenario_generation import sceneration
from datafev.routines.scenario_generation import utils
from datafev.data_handling.fleet import EVFleet
from datafev.data_handling.cluster import ChargerCluster
from datafev.data_handling.multi_cluster import MultiClusterSystem
from datafev.routines.arrival import *
from datafev.routines.departure import *
from datafev.routines.charging_control.decentralized_fcfs import charging_routine

import numpy as np
import pandas as pd


NUMBER_OF_EVs = 146
np.random.seed(0)


def step0_monkey_patch_openpyxl():
    """
    Some versions of openpyxl may crash for workbooks that come with wrong cell styling information.
    In that case, Monkey patch the library, so that it can still handle these types of workbooks.
    """
    from openpyxl.worksheet import _reader
    from openpyxl.cell import Cell

    def bind_cells(self):
        for idx, row in self.parser.parse():
            for cell in row:
                try:
                    style = self.ws.parent._cell_styles[cell['style_id']]
                except:  # This is the patch, original doesn't have a try/except here
                    style = None
                c = Cell(self.ws, row=cell['row'], column=cell['column'], style_array=style)
                c._value = cell['value']
                c.data_type = cell['data_type']
                self.ws._cells[(cell['row'], cell['column'])] = c
        self.ws.formula_attributes = self.parser.array_formulae
        if self.ws._cells:
            self.ws._current_row = self.ws.max_row  # use cells not row dimensions

    _reader.WorksheetReader.bind_cells = bind_cells
    return


def copy_cellstyle_openpyxl(new_cell, reference_cell):
    if reference_cell.has_style:
        new_cell.font = copy(reference_cell.font)
        new_cell.border = copy(reference_cell.border)
        new_cell.fill = copy(reference_cell.fill)
        new_cell.number_format = copy(reference_cell.number_format)
        new_cell.protection = copy(reference_cell.protection)
        new_cell.alignment = copy(reference_cell.alignment)
    return


def step1_generate_reference_fleet(statistical_input_file):
    """
    Example script for generating reference fleet behavior based on conditional pdfs.
    """

    (
        end_time,
        times_dict,
        times_prob_dict,
        soc_dict,
        soc_prob_dict,
        ev_dict,
    ) = utils.excel_to_sceneration_input_conditional_pdfs(
        file_path=statistical_input_file
    )

    success = False
    while not success:
        try:
            ev_df = sceneration.generate_fleet_from_conditional_pdfs(
                soc_dict=soc_dict,
                soc_prob_dict=soc_prob_dict,
                times_dict=times_dict,
                times_prob_dict=times_prob_dict,
                ev_dict=ev_dict,
                number_of_evs=NUMBER_OF_EVs,
                endtime=end_time,
                timedelta_in_min=15,
                diff_arr_dep_in_min=0,
            )
            success = True
        except:
            success = False

    # Unlocalize datetimes, as Excel does not support datetimes with timezones
    ev_df["ArrivalTime"] = ev_df["ArrivalTime"].dt.tz_localize(None)
    ev_df["DepartureTime"] = ev_df["DepartureTime"].dt.tz_localize(None)
    
    return ev_df


def step2_replicate_reference_fleet(step1_output, annual_horizon):
    """
    This function converts the fleet behavior (generated from statistical data) to the format that could be simulated
    in datafev simulations with longer simulation horizo.

    Parameters
    ----------
    step1_output : pandas.core.frame.DataFrame
        Output data frame from generate_fleet_data function.

        The target day to which the reference fleet behavior will be mapped.

    """

    fleet_annual = pd.DataFrame(
        columns=[
            "ev_id",
            "Battery Capacity (kWh)",
            "p_max_ch (kW)",
            "p_max_ds (kW)",
            "Reservation Time",
            "Estimated Arrival Time",
            "Estimated Departure Time",
            "Estimated Arrival SOC",
            "Target SOC @ Estimated Departure Time",
            "V2G Allowance (kWh)",
            "Real Arrival Time",
            "Real Arrival SOC",
            "Real Departure Time",
            "Target Cluster"
        ]
    )
    
    fleet_empty = fleet_annual.copy()

    for day in annual_horizon:
        fleet_daily = fleet_empty.copy()

        fleet_daily["Battery Capacity (kWh)"] = step1_output["BatteryCapacity(kWh)"].values
        fleet_daily["p_max_ch (kW)"] = step1_output["MaxFastChargingPower(kW)"].values
        fleet_daily["p_max_ds (kW)"] = 0.0
        
        # Arrival times, departure times, and arrival SOCs are shuffled every day
        fleet_daily["Real Arrival Time"] = (pd.to_datetime(step1_output["ArrivalTime"]).dt.time.apply(
            lambda t: pd.Timestamp.combine(day, t))).sample(frac=1).reset_index(drop=True)
        fleet_daily["Real Departure Time"] = (fleet_daily["Real Arrival Time"] +
                                              (step1_output["DepartureTime"]-step1_output["ArrivalTime"]))  # Parking
        fleet_daily["Real Arrival SOC"] = step1_output["ArrivalSoC"].sample(frac=1).reset_index(drop=True)  
        
        fleet_daily["ev_id"] = "v"+(fleet_daily.index + 1).map(lambda x: str(x).zfill(3))
        
        fleet_annual = pd.concat([fleet_annual, fleet_daily], ignore_index=True)

    return fleet_annual


def step3_charging_simulation(fleet_excel_file):
    # Simulation inputs
    input_file = pd.ExcelFile(fleet_excel_file)
    input_fleet = pd.read_excel(input_file, "Fleet", parse_dates=['Estimated Arrival Time',
                                                                  'Estimated Departure Time',
                                                                  'Real Arrival Time',
                                                                  'Real Departure Time'])

    input_cluster = np.empty(NUMBER_OF_EVs, dtype=object)
    input_capacity = np.empty(NUMBER_OF_EVs, dtype=object)

    for i in range(NUMBER_OF_EVs):
        cluster = pd.read_excel(input_file, "Cluster")
        cluster['cu_id'] = str("cluster" + str((i % NUMBER_OF_EVs)+1) + "")
        input_cluster[i] = cluster
        input_capacity[i] = pd.read_excel(input_file, "Capacity", parse_dates=['TimeStep'],
                                          date_format='%d.%m.%Y %H:%M:%S')

    # Getting the path of the input Excel file
    abs_path_input = os.path.abspath(input_file)
    print("Scenario inputs are taken from the xlsx file:", abs_path_input)
    print()

    # Simulation parameters
    sim_start = datetime(2019, 1, 1, 0, 0, 0)
    sim_end = datetime(2020, 1, 1, 23, 45, 0)
    sim_length = sim_end - sim_start
    sim_step = timedelta(minutes=15)
    sim_horizon = [sim_start + t * sim_step for t in range(int(sim_length / sim_step))]
    print("Simulation starts at:", sim_start)
    print("Simulation fininshes at:", sim_end)
    print("Length of one time step in simulation:", sim_step)
    print()

    # INITIALIZATION OF THE SIMULATION

    system = MultiClusterSystem("multicluster")
    for i in range(NUMBER_OF_EVs):
        print("Initializing simulation scenario " + str(i+1) + "/" + str(NUMBER_OF_EVs))
        cluster = ChargerCluster(str("cluster" + str((i % NUMBER_OF_EVs)+1) + ""), input_cluster[i])
        system.add_cc(cluster)
        fleet = EVFleet("fleet", input_fleet, sim_horizon)
        cluster.enter_power_limits(sim_start, sim_end, sim_step, input_capacity[i])

    print()
    print("All simulation scenario have been initialized. Starting the simulation.")
    print()

    # DYNAMIC SIMULATION

    for ts in sim_horizon:
        print("     Simulating time step:", ts)

        # The departure routine for the EVs leaving the charger clusters
        departure_routine(ts, fleet)

        # The arrival routine for the EVs incoming to the charger clusters
        arrival_routine(ts, sim_step, fleet, system)

        # Real-time charging control of the charger clusters
        charging_routine(ts, sim_step, system)

    print("Simulation finished...")
    print()
    print()

    # Printing the results to Excel files
    system.export_results_to_excel(
        sim_start, sim_end, sim_step, "Results_Cluster.xlsx"
    )
    fleet.export_results_to_excel(
        sim_start, sim_end, sim_step, "Results_Fleet.xlsx"
    )
    # Path of the output Excel file
    abs_path_output_cluster = os.path.abspath("Results_Cluster.xlsx")
    abs_path_output_fleet = os.path.abspath("Results_Fleet.xlsx")
    print("Scenario results are saved to the following xlsx files:")
    print(abs_path_output_cluster)
    print(abs_path_output_fleet)
    print()
    return


def step4_generate_csv_profiles(cluster_excel_file, output_path):
    wb = openpyxl.load_workbook(cluster_excel_file)
    ws = wb["Consumption (Aggregate)"]

    # Create the time samples:
    def datetime_range(start, end, delta):
        current = start
        while current < end:
            yield current
            current += delta

    dts = [dt.strftime('%Y-%m-%d %H:%M:00') for dt in datetime_range(
        datetime(2021, 1, 1, 0, 0),
        datetime(2021, 12, 31, 23, 59),
        timedelta(minutes=15))]

    number_ev = len(ws["1"]) - 1
    timeslots = len(dts)

    for i in range(number_ev):
        csv_file_name = "ev_" + str(i) + ".csv"
        print("Writing CSV file " + str(csv_file_name) + ".")

        ev_profile = np.empty(timeslots, dtype=float)
        for j in range(timeslots):
            ev_profile[j] = float(ws.cell(j + 2, i + 2).value)

        # We don't want to charge an EV every single day. Target consumption should be approx. 2.500kWh ==> 15.000km.
        randint = np.random.choice(range(364), 275, replace=False)
        for x in randint:
            for j in range(96):
                ev_profile[x * 96 + j] = 0.0

        # Save the profile:
        with open(os.path.join(output_path, csv_file_name), 'w', newline='') as file:
            writer = csv.writer(file, delimiter=",", escapechar='', quoting=csv.QUOTE_NONE)
            writer.writerow(["YYYY-MM-DD HH:MM:SS", "Ladestrom [kW]"])
            for j in range(len(ev_profile)):
                s = dts[j]
                d = s[:10]
                ti = s[-8:]
                writer.writerow([d + " " + ti, str(ev_profile[j])])

    csv_file_name = "no_ev.csv"
    print("Writing CSV file " + str(csv_file_name) + ".")

    ev_profile = np.zeros(timeslots, dtype=float)

    # Save the profile:
    with open(os.path.join(output_path, csv_file_name), 'w', newline='') as file:
        writer = csv.writer(file, delimiter=",", escapechar='', quoting=csv.QUOTE_NONE)
        writer.writerow(["YYYY-MM-DD HH:MM:SS", "Ladestrom [kW]"])
        for j in range(len(ev_profile)):
            s = dts[j]
            d = s[:10]
            ti = s[-8:]
            writer.writerow([d + " " + ti, str(ev_profile[j])])


if __name__ == '__main__':
    #step0_monkey_patch_openpyxl()

    print("Running step no. 1 ...")
    # Statistics: https://www.mobilitaet-in-deutschland.de/archive/publikationen2017.html
    ref_fleet = step1_generate_reference_fleet("Statistics.xlsx")

    annnual_horizon = pd.date_range(start="2019-01-01", freq="1D", periods=365)

    print("Running step no. 2 ...")
    fleet = step2_replicate_reference_fleet(ref_fleet, annnual_horizon)
    fleet.to_excel("Fleet.xlsx")

    time.sleep(2.0)
    print("Manipulating input file Fleet.xlsx using openpyxl ...")

    wb = openpyxl.load_workbook("Fleet.xlsx")
    ws0 = wb.worksheets[0]
    ws0.title = u'Fleet'

    for i in range(len(ws0['B'])-1):

        ws0.cell(row=i+2, column=7).value = ws0.cell(row=i+2, column=12).value
        copy_cellstyle_openpyxl(ws0.cell(row=i+2, column=7), ws0['L2'])
        ws0.cell(row=i+2, column=8).value = ws0.cell(row=i+2, column=14).value
        copy_cellstyle_openpyxl(ws0.cell(row=i+2, column=8), ws0['N2'])
        ws0.cell(row=i+2, column=10).value = float(1.0)
        copy_cellstyle_openpyxl(ws0.cell(row=i+2, column=10), ws0['M2'])
        ws0.cell(row=i+2, column=15).value = str("cluster" + str((i % NUMBER_OF_EVs)+1) + "")
        copy_cellstyle_openpyxl(ws0.cell(row=i+2, column=15), ws0['B2'])

    ws1 = wb.create_sheet("Cluster")
    ws1['A1'].value = 'cu_id'
    copy_cellstyle_openpyxl(ws1['A1'], ws0['B1'])
    ws1['B1'].value = 'cu_p_ch_max (kW)'
    copy_cellstyle_openpyxl(ws1['B1'], ws0['B1'])
    ws1['C1'].value = 'cu_p_ds_max (kW)'
    copy_cellstyle_openpyxl(ws1['C1'], ws0['B1'])
    ws1['D1'].value = 'cu_eff'
    copy_cellstyle_openpyxl(ws1['D1'], ws0['B1'])

    ws1['A2'].value = 'CC'
    ws1['B2'].value = float(11)
    ws1['C2'].value = float(0)
    ws1['D2'].value = float(1)

    ws2 = wb.create_sheet("Capacity")
    ws2['A1'].value = 'TimeStep'
    copy_cellstyle_openpyxl(ws2['A1'], ws0['B1'])
    ws2['B1'].value = 'LB (kW)'
    copy_cellstyle_openpyxl(ws2['B1'], ws0['B1'])
    ws2['C1'].value = 'UB (kW)'
    copy_cellstyle_openpyxl(ws2['C1'], ws0['B1'])

    annnual_horizon = pd.date_range(start="2019-01-01", freq="0.25H", periods=365*96)
    for t in range(len(annnual_horizon)):
        ws2.cell(row=t+2, column=1).value = str(annnual_horizon[t].strftime('%d.%m.%Y %H:%M:%S'))
        copy_cellstyle_openpyxl(ws2.cell(row=t+2, column=1), ws0['L2'])
        ws2.cell(row=t+2, column=2).value = float(0.0)
        ws2.cell(row=t+2, column=3).value = float(99999.9)

    wb.save("Fleet.xlsx")

    print("Done.")
    time.sleep(2.0)

    print("Running step no. 3 ...")
    step3_charging_simulation("Fleet.xlsx")

    print("Running step no. 4 ...")
    step4_generate_csv_profiles(
        "Results_Cluster.xlsx",
        os.path.abspath(os.path.join(os.path.dirname(os.path.abspath(sys.argv[0])), "profiles"))
    )
